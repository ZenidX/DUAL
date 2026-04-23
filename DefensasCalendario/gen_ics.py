# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, timedelta
import re

EXCEL_PATH = 'd:/Xavi/ProjectsITIC/DUAL-SolucionesClaude/DefensasCalendario/Plantilla_MG2025-2026ITICBCN Avaluació defensa Memòria Dossier DUAL-Complecio.xlsx'
OUTPUT_PATH = 'd:/Xavi/ProjectsITIC/DUAL-SolucionesClaude/DefensasCalendario/Calendario.ics'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Horaris 25-26']


def fold_line(line):
    """RFC 5545: lines longer than 75 octets must be folded"""
    encoded = line.encode('utf-8')
    if len(encoded) <= 75:
        return line
    result = []
    current = ''
    for char in line:
        test = (current + char).encode('utf-8')
        if len(test) > 75:
            result.append(current)
            current = ' ' + char
        else:
            current += char
    if current:
        result.append(current)
    return '\r\n'.join(result)


def escape_ical(text):
    if not text:
        return ''
    text = text.replace('\\', '\\\\')
    text = text.replace(';', '\\;')
    text = text.replace(',', '\\,')
    text = text.replace('\n', '\\n')
    return text


DESCRIPTION_TEMPLATE = (
    "CONVOCATORIA DE DEFENSA DEL PROJECTE DUAL 2025-2026\\n"
    "\\n"
    "Alumne: {alumne}\\n"
    "Classe: {classe}\\n"
    "Tutor: {tutor}\\n"
    "Tribunal: {coord1} / {coord2}\\n"
    "\\n"
    "--- FUNCIONAMENT DE LA DEFENSA ---\\n"
    "\\n"
    "1. PRESENTACIO (10 min):\\n"
    "   L'alumne exposa el projecte realitzat a l'empresa durant el periode de DUAL. "
    "Cal explicar els objectius\\, les tasques realitzades\\, les tecnologies utilitzades\\, "
    "els resultats obtinguts i les competencies adquirides.\\n"
    "\\n"
    "2. TORN DE PREGUNTES (5 min):\\n"
    "   El tribunal (coordinadors i tutor) formulara preguntes sobre el projecte\\, "
    "el dossier i l'experiencia a l'empresa. "
    "L'alumne haura de defensar i justificar les decisions preses.\\n"
    "\\n"
    "--- DOCUMENTACIO REQUERIDA ---\\n"
    "\\n"
    "- Dossier de seguiment\\n"
    "- Presentacio (PowerPoint\\, Google Slides o similar)\\n"
    "\\n"
    "Duracio total: 20 minuts\\n"
    "Lloc: ITIC Barcelona (o videoconferencia si s'indica)"
)

events = []

for row_idx in range(3, ws.max_row + 1):
    row = ws[row_idx]
    date_val = row[0].value
    classe = row[1].value
    alumne = row[2].value
    coord1 = row[3].value
    coord2 = row[4].value
    tutor = row[5].value
    email_al = row[8].value
    email_c1 = row[9].value
    email_c2 = row[10].value
    email_tutor = row[11].value

    if not alumne or not date_val:
        continue

    alumne = alumne.strip()
    classe = (classe or '').strip()
    coord1 = (coord1 or '').strip()
    coord2 = (coord2 or '').strip()
    tutor = (tutor or '').strip()

    if not isinstance(date_val, datetime):
        continue

    # Skip students with problems (missing email alumno or tutor)
    if not email_al or not email_tutor:
        print(f'  SKIPPED (missing data): {alumne}')
        continue

    dt_start = date_val
    dt_end = dt_start + timedelta(minutes=20)

    dtstart_str = dt_start.strftime('%Y%m%dT%H%M%S')
    dtend_str = dt_end.strftime('%Y%m%dT%H%M%S')

    summary = f'{alumne} ({classe}) - Defensa Projecte DUAL'
    # SUMMARY: no escaping needed for display text in Google Calendar
    summary_ical = summary.replace(',', '\\,').replace(';', '\\;')

    desc = DESCRIPTION_TEMPLATE.format(
        alumne=escape_ical(alumne),
        classe=escape_ical(classe),
        tutor=escape_ical(tutor),
        coord1=escape_ical(coord1),
        coord2=escape_ical(coord2)
    )

    slug = re.sub(r'[^a-z0-9]', '-', alumne.lower())
    uid = f'{dtstart_str}-defensa-{slug}@iticbcn.cat'

    def quote_cn(name):
        """Quote CN value if it contains special chars"""
        if ',' in name or ';' in name or ':' in name:
            return f'"{name}"'
        return name

    attendees = []
    if email_al:
        attendees.append(
            f'ATTENDEE;CN={quote_cn(alumne)};ROLE=REQ-PARTICIPANT;RSVP=TRUE:MAILTO:{email_al.strip()}'
        )
    if email_tutor and email_tutor.strip():
        attendees.append(
            f'ATTENDEE;CN={quote_cn(tutor)};ROLE=REQ-PARTICIPANT;RSVP=TRUE:MAILTO:{email_tutor.strip()}'
        )
    if email_c1 and email_c1.strip():
        attendees.append(
            f'ATTENDEE;CN={quote_cn(coord1)};ROLE=REQ-PARTICIPANT;RSVP=TRUE:MAILTO:{email_c1.strip()}'
        )
    if email_c2 and email_c2.strip():
        if email_c2.strip() != (email_tutor or '').strip():
            attendees.append(
                f'ATTENDEE;CN={quote_cn(coord2)};ROLE=REQ-PARTICIPANT;RSVP=TRUE:MAILTO:{email_c2.strip()}'
            )

    event_lines = [
        'BEGIN:VEVENT',
        f'DTSTART;TZID=Europe/Madrid:{dtstart_str}',
        f'DTEND;TZID=Europe/Madrid:{dtend_str}',
        f'SUMMARY:{summary_ical}',
        f'DESCRIPTION:{desc}',
        f'ORGANIZER;CN=ITIC Barcelona:MAILTO:xavi.lara@iticbcn.cat',
    ]
    for att in attendees:
        event_lines.append(att)
    event_lines.extend([
        'STATUS:CONFIRMED',
        f'UID:{uid}',
        'END:VEVENT',
    ])
    events.append(event_lines)

# Build full ICS
VTIMEZONE = [
    'BEGIN:VTIMEZONE',
    'TZID:Europe/Madrid',
    'BEGIN:STANDARD',
    'DTSTART:19701025T030000',
    'RRULE:FREQ=YEARLY;BYDAY=-1SU;BYMONTH=10',
    'TZOFFSETFROM:+0200',
    'TZOFFSETTO:+0100',
    'TZNAME:CET',
    'END:STANDARD',
    'BEGIN:DAYLIGHT',
    'DTSTART:19700329T020000',
    'RRULE:FREQ=YEARLY;BYDAY=-1SU;BYMONTH=3',
    'TZOFFSETFROM:+0100',
    'TZOFFSETTO:+0200',
    'TZNAME:CEST',
    'END:DAYLIGHT',
    'END:VTIMEZONE',
]

header = [
    'BEGIN:VCALENDAR',
    'VERSION:2.0',
    'PRODID:-//ITIC BCN//Defensas Projecte DUAL 2025-2026//CA',
    'CALSCALE:GREGORIAN',
    'METHOD:PUBLISH',
    'X-WR-CALNAME:Defensas Projecte DUAL ITIC BCN',
    'X-WR-TIMEZONE:Europe/Madrid',
]

all_lines = header + VTIMEZONE
for event in events:
    all_lines.extend(event)
all_lines.append('END:VCALENDAR')

# Fold long lines and write
output_lines = []
for line in all_lines:
    output_lines.append(fold_line(line))

content = '\r\n'.join(output_lines)

with open(OUTPUT_PATH, 'wb') as f:
    f.write(content.encode('utf-8'))

print(f'Generated {len(events)} events')
for ev in events[:3]:
    for line in ev:
        if line.startswith('SUMMARY:') or line.startswith('ATTENDEE'):
            print(f'  {line}')
    print()
